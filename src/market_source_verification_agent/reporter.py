"""Report rendering for verified claims."""

from __future__ import annotations

import html
import io
import json
import re
from typing import Literal

from .schema import Claim, ClassifyResult, IR, VerifyResult

ReportFormat = Literal["xlsx", "md", "html", "json"]

VERDICT_LABELS = {
    "supported": "✅ 支持",
    "partially_supported": "⚠️ 部分支持",
    "not_found": "❌ 未找到",
    "contradicted": "❗ 矛盾",
    "not_verifiable": "❓ 无法验证",
}

FILL_COLORS = {
    "supported": "C6EFCE",
    "partially_supported": "FFEB9C",
    "not_found": "D9D9D9",
    "contradicted": "FFC7CE",
    "not_verifiable": "BFBFBF",
}


def render(
    ir: IR,
    claims: list[Claim],
    verifies: dict[str, VerifyResult],
    classes: dict[str, ClassifyResult],
    fmt: ReportFormat = "xlsx",
    detailed: bool = False,
) -> bytes:
    # 新格式（ClaimID 表）走分块输出：每个 sub-table 各自的列头 + 来源是否真实/类别
    new_format_claims = [c for c in claims if c.original_columns]
    if new_format_claims and len(new_format_claims) == len(claims):
        sections = _report_sections(claims, verifies, classes, detailed)
        if fmt == "json":
            return json.dumps([
                {"title": s["title"], "headers": s["headers"], "rows": s["rows"]}
                for s in sections
            ], ensure_ascii=False, indent=2).encode("utf-8")
        if fmt == "md":
            return _render_md_sections(sections, detailed).encode("utf-8")
        if fmt == "html":
            return _render_html_sections(sections, detailed).encode("utf-8")
        return _render_xlsx_sections(sections, detailed)

    # 兼容旧格式（数值表 / 段落 claim）
    rows = _report_rows(claims, verifies, classes, detailed)
    if fmt == "json":
        return json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8")
    if fmt == "md":
        return _render_md(rows, detailed).encode("utf-8")
    if fmt == "html":
        return _render_html(rows, detailed).encode("utf-8")
    return _render_xlsx(rows, detailed)


def _report_sections(
    claims: list[Claim],
    verifies: dict[str, VerifyResult],
    classes: dict[str, ClassifyResult],
    detailed: bool,
) -> list[dict]:
    """按 sub-table 分组：同 table_signature 的 claim 一组。每组各自列头 + 追加 2 列。"""
    grouped: dict[str, list[Claim]] = {}
    order: list[str] = []
    for c in claims:
        sig = c.table_signature or "_default"
        if sig not in grouped:
            grouped[sig] = []
            order.append(sig)
        grouped[sig].append(c)

    sections = []
    for sig in order:
        group = grouped[sig]
        # 列名：保持首条 claim 的 original_columns 顺序（同 sub-table 顺序一致）
        first_oc = group[0].original_columns or {}
        col_names: list[str] = list(first_oc.keys())
        # 合并补齐：以防某些 claim 多出新列
        for c in group[1:]:
            for k in (c.original_columns or {}):
                if k not in col_names:
                    col_names.append(k)
        appended = ["来源是否真实", "原因", "核验诊断", "多源核验明细", "来源类别"]
        if detailed:
            appended.append("核验佐证")
        headers = col_names + appended

        rows = []
        for c in group:
            verify = verifies.get(c.claim_id)
            category = classes.get(c.claim_id)
            row: dict[str, str] = {}
            for col in col_names:
                row[col] = (c.original_columns or {}).get(col, "")
            row["原始ClaimID"] = c.original_claim_id or c.claim_id
            row["来源名称"] = c.source_name_raw
            row["来源URL提示"] = c.source_url_hint or ""
            row["来源URL列表"] = "\n".join(c.source_urls or [url for url in [c.source_url_hint, *c.extra_source_urls] if url])
            row["来源是否真实"] = VERDICT_LABELS[verify.verdict] if verify else "❓ 无法验证"
            row["原因"] = _diagnosis_summary(verify) if verify else ""
            row["核验诊断"] = _diagnosis_summary(verify) if verify else ""
            row["多源核验明细"] = _source_details_text(verify) if verify else ""
            row["来源类别"] = category.tier if category else ""
            if detailed:
                row["核验佐证"] = _detail_text(verify, category) if verify and category else ""
            rows.append(row)

        # 节标题：用 sub-table 列签名前 N 个非空列名拼出
        title = " / ".join([n for n in col_names if n][:4])
        sections.append({"title": title or "sub-table", "headers": headers, "rows": rows})
    return sections


def _render_xlsx_sections(sections: list[dict], detailed: bool) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("xlsx output requires openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "source_verification"
    section_title_font = Font(bold=True, size=12)

    for section in sections:
        # Section title row
        ws.append([f"【{section['title']}】"])
        ws.cell(row=ws.max_row, column=1).font = section_title_font
        # Header row
        ws.append(section["headers"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        # Data rows
        for row in section["rows"]:
            values = [row.get(h, "") for h in section["headers"]]
            ws.append(values)
            verdict_idx = section["headers"].index("来源是否真实") + 1
            verdict_cell = ws.cell(row=ws.max_row, column=verdict_idx)
            verdict_key = _verdict_key(str(verdict_cell.value))
            verdict_cell.fill = PatternFill(fill_type="solid", fgColor=FILL_COLORS.get(verdict_key, "FFFFFF"))
        ws.append([])  # 空行分隔

    # 列宽合理默认
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = 24

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _render_md_sections(sections: list[dict], detailed: bool) -> str:
    out = []
    for section in sections:
        out.append(f"## {section['title']}\n")
        headers = section["headers"]
        out.append("| " + " | ".join(headers) + " |")
        out.append("| " + " | ".join("---" for _ in headers) + " |")
        for row in section["rows"]:
            out.append("| " + " | ".join(_md_cell(row.get(h, "")) for h in headers) + " |")
        out.append("")
    return "\n".join(out) + "\n"


def _render_html_sections(sections: list[dict], detailed: bool) -> str:
    body_parts = []
    for section in sections:
        body_parts.append(f"<h2>{html.escape(section['title'])}</h2>")
        head = "".join(f"<th>{html.escape(h)}</th>" for h in section["headers"])
        rows_html = []
        for row in section["rows"]:
            cells = "".join(f"<td>{html.escape(str(row.get(h, '') or ''))}</td>" for h in section["headers"])
            rows_html.append(f"<tr>{cells}</tr>")
        body_parts.append(
            f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(rows_html)}</tbody></table>"
        )
    return f"""<!doctype html>
<html lang="zh-CN">
<meta charset="utf-8">
<title>来源核验报告</title>
<style>
body{{font-family:Arial,"Microsoft YaHei",sans-serif;margin:24px;color:#17202a}}
h2{{margin-top:32px;color:#1f3a5f}}
table{{border-collapse:collapse;width:100%;font-size:14px;margin-bottom:24px}}
th,td{{border:1px solid #d7dde5;padding:8px;vertical-align:top}}
th{{background:#f4f6f8;text-align:left}}
tr:nth-child(even){{background:#fbfcfd}}
</style>
{''.join(body_parts)}
</html>"""


def summarize(verifies: dict[str, VerifyResult], classes: dict[str, ClassifyResult]) -> dict[str, int]:
    summary: dict[str, int] = {"total": len(verifies)}
    for result in verifies.values():
        summary[result.verdict] = summary.get(result.verdict, 0) + 1
    for result in classes.values():
        summary[result.tier] = summary.get(result.tier, 0) + 1
    return summary


def detailed_summary(
    verifies: dict[str, VerifyResult],
    classes: dict[str, ClassifyResult],
) -> dict[str, object]:
    """Verdict / tier counts plus per-domain failure rate, suitable for a sidecar summary file."""
    base: dict[str, object] = dict(summarize(verifies, classes))

    domain_total: dict[str, int] = {}
    domain_fail: dict[str, int] = {}
    for verify in verifies.values():
        for detail in verify.source_details or []:
            domain = str(detail.get("domain") or "unknown")
            status = str(detail.get("fetch_status") or "")
            verdict = str(detail.get("verdict") or "")
            domain_total[domain] = domain_total.get(domain, 0) + 1
            if status != "ok" or verdict in {"not_found", "not_verifiable"}:
                domain_fail[domain] = domain_fail.get(domain, 0) + 1

    by_domain: dict[str, dict[str, int | float]] = {}
    for domain, total in domain_total.items():
        fail = domain_fail.get(domain, 0)
        by_domain[domain] = {
            "total": total,
            "fail": fail,
            "fail_rate": round(fail / total, 3) if total else 0.0,
        }
    base["by_domain"] = by_domain
    try:
        from .resolver import get_llm_extract_calls

        base["llm_extract_calls"] = get_llm_extract_calls()
    except Exception:
        pass
    return base


def aggregate_multisource_verdict(
    source_details: list[dict],
) -> tuple[str, str]:
    """Derive a final verdict from per-source details using the documented policy.

    Returns (verdict, reason). Verdict is one of supported/partially_supported/
    contradicted/not_found/not_verifiable.
    """
    if not source_details:
        return "not_verifiable", "no source details"

    verdicts = [str(d.get("verdict") or "") for d in source_details]
    statuses = [str(d.get("fetch_status") or "") for d in source_details]
    fetch_ok = [s == "ok" for s in statuses]

    if "supported" in verdicts:
        return "supported", "at least one source supports the claim"
    if "contradicted" in verdicts:
        return "contradicted", "at least one source contradicts the claim"
    if "partially_supported" in verdicts:
        return "partially_supported", "at least one source partially supports the claim"
    if any(fetch_ok) and "not_found" in verdicts:
        return "not_found", "all reachable sources lack the claim keywords"
    return "not_verifiable", "all sources failed to fetch or were unreadable"


def _report_rows(
    claims: list[Claim],
    verifies: dict[str, VerifyResult],
    classes: dict[str, ClassifyResult],
    detailed: bool,
) -> list[dict[str, str | None]]:
    rows: list[dict[str, str | None]] = []
    for claim in claims:
        verify = verifies[claim.claim_id]
        category = classes[claim.claim_id]
        row: dict[str, str | None] = {
            "章节": " / ".join(claim.section_path),
            "指标": claim.metric,
            "数值": claim.value,
            "年份": claim.year,
            "地区/口径": claim.region,
            "事实声明": claim.statement,
            "来源名称": claim.source_name_raw,
            "来源URL提示": claim.source_url_hint,
            "来源URL列表": "\n".join(claim.source_urls or [url for url in [claim.source_url_hint, *claim.extra_source_urls] if url]),
            "重复条数": str(claim.duplicate_count),
            "原始Claim列表": "\n".join(claim.duplicate_claim_ids),
            "核验诊断": _diagnosis_summary(verify),
            "多源核验明细": _source_details_text(verify),
            "来源是否真实": VERDICT_LABELS[verify.verdict],
            "来源类别": category.tier,
        }
        if detailed:
            row["核验佐证"] = _detail_text(verify, category)
        rows.append(row)
    return rows


def _render_xlsx(rows: list[dict[str, str | None]], detailed: bool) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("xlsx output requires openpyxl; install project dependencies first") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "source_verification"
    headers = _headers(detailed)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(header) for header in headers])
        verdict_cell = ws.cell(row=ws.max_row, column=headers.index("来源是否真实") + 1)
        verdict_key = _verdict_key(str(verdict_cell.value))
        verdict_cell.fill = PatternFill(fill_type="solid", fgColor=FILL_COLORS.get(verdict_key, "FFFFFF"))

    widths = {"章节": 24, "指标": 24, "数值": 18, "年份": 14, "地区/口径": 18, "事实声明": 42, "来源名称": 34, "来源URL提示": 30, "来源URL列表": 42, "重复条数": 10, "原始Claim列表": 32, "核验诊断": 36, "多源核验明细": 72, "来源是否真实": 16, "来源类别": 10, "核验佐证": 60}
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(header, 18)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _render_md(rows: list[dict[str, str | None]], detailed: bool) -> str:
    headers = _headers(detailed)
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    for row in rows:
        lines.append("| " + " | ".join(_md_cell(row.get(header)) for header in headers) + " |")
    return "\n".join(lines) + "\n"


def _render_html(rows: list[dict[str, str | None]], detailed: bool) -> str:
    headers = _headers(detailed)
    head = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
    body = []
    for row in rows:
        cells = "".join(f"<td>{html.escape(str(row.get(header) or ''))}</td>" for header in headers)
        body.append(f"<tr>{cells}</tr>")
    return f"""<!doctype html>
<html lang="zh-CN">
<meta charset="utf-8">
<title>来源核验报告</title>
<style>
body{{font-family:Arial,"Microsoft YaHei",sans-serif;margin:24px;color:#17202a}}
table{{border-collapse:collapse;width:100%;font-size:14px}}
th,td{{border:1px solid #d7dde5;padding:8px;vertical-align:top}}
th{{background:#f4f6f8;text-align:left}}
tr:nth-child(even){{background:#fbfcfd}}
</style>
<table><thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table>
</html>"""


def _headers(detailed: bool) -> list[str]:
    headers = ["章节", "指标", "数值", "年份", "地区/口径", "事实声明", "来源名称", "来源URL提示", "来源URL列表", "重复条数", "原始Claim列表", "核验诊断", "多源核验明细", "来源是否真实", "来源类别"]
    if detailed:
        headers.append("核验佐证")
    return headers


def _detail_text(verify: VerifyResult, category: ClassifyResult) -> str:
    if verify.verdict in {"not_found", "not_verifiable"}:
        return _diagnosis_summary(verify)
    parts = [
        _clean_detail_piece(verify.evidence_quote),
        verify.evidence_locator,
        verify.discrepancy,
        _clean_detail_piece(verify.reasoning),
        category.tier_reason,
    ]
    return "；".join(part for part in parts if part)


def _source_details_text(verify: VerifyResult) -> str:
    if not verify.source_details:
        return ""
    lines = []
    for idx, detail in enumerate(verify.source_details, start=1):
        url = detail.get("url") or ""
        verdict = detail.get("verdict") or ""
        tier = detail.get("tier") or ""
        status = detail.get("fetch_status") or ""
        content_type = detail.get("content_type") or ""
        method = detail.get("resolution_method") or ""
        diagnosis = _diagnose_detail(detail)
        reason = detail.get("reasoning") or ""
        quote = detail.get("evidence_quote") or ""
        if verdict in {"not_found", "not_verifiable"}:
            reason = ""
            quote = ""
        else:
            reason = _clean_detail_piece(str(reason))
            quote = _clean_detail_piece(str(quote))
        lines.append(
            f"{idx}. [{tier}/{verdict}/{status}/{content_type}/{method}] {url}\n"
            f"诊断：{diagnosis}\n"
            f"{quote}\n"
            f"{reason}".strip()
        )
    return "\n\n".join(lines)


def _clean_detail_piece(value: str | None) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    compact = re.sub(r"\s+", "", text)
    signal_count = sum(1 for token in ("#", "$", "{", "}", "_", "&amp;", "锛", "鈥") if token in text)
    cjk_count = len(re.findall(r"[\u4e00-\u9fff]", text))
    if len(compact) > 40 and signal_count >= 3 and cjk_count < len(compact) * 0.25:
        return ""
    return text


def _diagnosis_summary(verify: VerifyResult) -> str:
    if not verify.source_details:
        return _diagnose_reasoning(verify.verdict, "", "", "", verify.reasoning)

    diagnoses = [_diagnose_detail(detail) for detail in verify.source_details]
    unique = list(dict.fromkeys(item for item in diagnoses if item))
    if not unique:
        return _diagnose_reasoning(verify.verdict, "", "", "", verify.reasoning)
    if len(unique) == 1:
        return unique[0]

    has_unreadable_pdf = any("PDF已下载" in item for item in unique)
    has_not_found = any("命中声明关键词" in item for item in unique)
    has_fetch_failure = any("链接无法访问" in item for item in unique)
    if has_fetch_failure and has_unreadable_pdf and has_not_found:
        return "部分链接无法访问；部分PDF已下载但文本不可读；部分来源已读取但未命中声明关键词"
    if has_unreadable_pdf and has_not_found:
        return "部分PDF已下载但文本不可读；部分来源已读取但未命中声明关键词"
    if has_fetch_failure and has_not_found:
        return "部分链接无法访问；部分来源已读取但未命中声明关键词"
    if has_fetch_failure and has_unreadable_pdf:
        return "部分链接无法访问；部分PDF已下载但文本不可读"
    return "；".join(unique[:3])


def _diagnose_detail(detail: dict[str, str | float | None]) -> str:
    verdict = str(detail.get("verdict") or "")
    status = str(detail.get("fetch_status") or "")
    content_type = str(detail.get("content_type") or "")
    method = str(detail.get("resolution_method") or "")
    reason = str(detail.get("reasoning") or "")
    return _diagnose_reasoning(verdict, status, content_type, method, reason)


_PDF_DIAG_RE = re.compile(r"\[diagnostic:\s*(.*?)\]\s*$")


def _extract_pdf_diagnostic(reason: str) -> str:
    match = _PDF_DIAG_RE.search(reason)
    if not match:
        return ""
    raw = match.group(1).strip()
    if "not a PDF" in raw:
        return "下载到的不是 PDF（疑似 HTML 错误页）"
    if "scanned/image-only" in raw or "all" in raw and "pages returned empty text" in raw:
        return "PDF 无文本层，疑为扫描件；OCR 未启用"
    if "pdfplumber.open failed" in raw:
        return f"PDF 解析失败：{raw}"
    if "cache file is empty" in raw:
        return "缓存文件为 0 字节"
    return raw


def _diagnose_reasoning(verdict: str, status: str, content_type: str, method: str, reason: str) -> str:
    if method == "failed" or status == "skipped":
        return "没有可解析的来源URL"
    if status and status != "ok":
        if status == "empty_body":
            return "服务器返回空响应（疑似反爬或登录重定向），无法核验"
        return f"链接无法访问：{status}"
    if status == "ok" and (
        "source content is empty or unreadable" in reason
        or "empty or unreadable" in reason
        or "text extraction returned empty" in reason
    ):
        diag = _extract_pdf_diagnostic(reason)
        if content_type == "pdf":
            base = "PDF已下载，但文本提取为空"
            if diag:
                return f"{base}（{diag}）"
            return f"{base}，可能是扫描件或不可读PDF；当前系统未启用OCR"
        return "页面可访问，但正文提取为空"
    if verdict == "not_found" or "no source passage mentions the claim keywords" in reason:
        return "来源已读取，但没有命中声明关键词"
    if verdict == "not_verifiable":
        return "来源可访问性或正文可读性不足，暂时无法验证"
    if verdict == "supported":
        return "来源已读取，并找到支持声明的证据"
    if verdict == "partially_supported":
        return "来源已读取，但只找到部分支持证据"
    if verdict == "contradicted":
        return "来源已读取，但内容与声明存在冲突"
    return reason or "暂无诊断"


def _md_cell(value: str | None) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", "<br>")


def _verdict_key(label: str) -> str:
    for key, value in VERDICT_LABELS.items():
        if value == label:
            return key
    return ""
