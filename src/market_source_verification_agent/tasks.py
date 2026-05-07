"""Local task service used by the FastAPI layer.

The production design in docs/07_orchestrator.md targets MongoDB + Redis/RQ.
This module keeps the same RunTask/Event/Artifact contracts while using the
filesystem as a development backend, so the API is usable before stage-three
infrastructure is provisioned.
"""

from __future__ import annotations

import hashlib
import json
import logging
import mimetypes
import shutil
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Protocol
from uuid import uuid4

from .config import ROOT, Settings, load_settings
from .orchestrator import run
from .schema import Artifact, RunEvent, RunTask
from .usage import UsageAccumulator, add_persistent_usage, compact_usage_summary, read_persistent_usage

logger = logging.getLogger(__name__)

CONTENT_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "md": "text/markdown; charset=utf-8",
    "html": "text/html; charset=utf-8",
    "json": "application/json; charset=utf-8",
}


class TaskStore(Protocol):
    settings: Settings

    def create_run(
        self,
        owner_id: str,
        input_kind: str,
        requested_format: str,
        detailed: bool,
        input_filename: str | None = None,
        input_path: str | None = None,
    ) -> RunTask: ...

    def get_task(self, run_id: str, owner_id: str) -> RunTask | None: ...

    def save_task(self, task: RunTask) -> None: ...

    def add_event(
        self,
        run_id: str,
        stage: str,
        level: str,
        message: str,
        progress_current: int | None = None,
        progress_total: int | None = None,
    ) -> RunEvent: ...

    def list_events(self, run_id: str, owner_id: str) -> list[RunEvent]: ...

    def save_artifact(self, artifact: Artifact) -> None: ...

    def list_artifacts(self, run_id: str, owner_id: str) -> list[Artifact]: ...

    def list_runs(self, owner_id: str, limit: int = 20, offset: int = 0) -> list[RunTask]: ...

    def delete_run(self, run_id: str, owner_id: str) -> bool: ...

    def delete_runs(self, owner_id: str) -> int: ...


class LocalTaskStore:
    """Tiny JSON-backed task store for local development and tests."""

    def __init__(self, settings: Settings | None = None):
        self.settings = settings or load_settings()
        self.root = _abs_path(self.settings.storage.reports_dir).parent / "runs"
        self.root.mkdir(parents=True, exist_ok=True)
        self._lock = threading.RLock()

    def create_run(
        self,
        owner_id: str,
        input_kind: str,
        requested_format: str,
        detailed: bool,
        input_filename: str | None = None,
        input_path: str | None = None,
    ) -> RunTask:
        now = datetime.now()
        task = RunTask(
            run_id=str(uuid4()),
            owner_id=owner_id,
            status="queued",
            input_kind=input_kind,  # type: ignore[arg-type]
            input_filename=input_filename,
            input_path=input_path,
            requested_format=requested_format,  # type: ignore[arg-type]
            detailed=detailed,
            created_at=now,
            updated_at=now,
        )
        self.save_task(task)
        self.add_event(task.run_id, "queued", "info", "Run queued")
        return task

    def get_task(self, run_id: str, owner_id: str) -> RunTask | None:
        path = self._task_path(run_id)
        if not path.exists():
            return None
        data = _read_json_text(path)
        if data is None:
            return None
        try:
            task = RunTask.model_validate_json(data)
        except Exception:
            return None
        return task if task.owner_id == owner_id else None

    def save_task(self, task: RunTask) -> None:
        task.updated_at = datetime.now()
        run_dir = self._run_dir(task.run_id)
        run_dir.mkdir(parents=True, exist_ok=True)
        with self._lock:
            _atomic_write_text(self._task_path(task.run_id), task.model_dump_json(indent=2))

    def add_event(
        self,
        run_id: str,
        stage: str,
        level: str,
        message: str,
        progress_current: int | None = None,
        progress_total: int | None = None,
    ) -> RunEvent:
        event = RunEvent(
            event_id=str(uuid4()),
            run_id=run_id,
            stage=stage,
            level=level,  # type: ignore[arg-type]
            message=message,
            progress_current=progress_current,
            progress_total=progress_total,
            created_at=datetime.now(),
        )
        path = self._events_path(run_id)
        with self._lock:
            events = _read_json_list(path)
            events.append(event.model_dump(mode="json"))
            _atomic_write_text(path, json.dumps(events, ensure_ascii=False, indent=2))
        return event

    def list_events(self, run_id: str, owner_id: str) -> list[RunEvent]:
        if not self.get_task(run_id, owner_id):
            return []
        path = self._events_path(run_id)
        if not path.exists():
            return []
        return [RunEvent.model_validate(item) for item in _read_json_list(path)]

    def save_artifact(self, artifact: Artifact) -> None:
        path = self._artifacts_path(artifact.run_id)
        with self._lock:
            artifacts = _read_json_list(path)
            artifacts = [item for item in artifacts if item.get("artifact_id") != artifact.artifact_id]
            artifacts.append(artifact.model_dump(mode="json"))
            _atomic_write_text(path, json.dumps(artifacts, ensure_ascii=False, indent=2))

    def list_artifacts(self, run_id: str, owner_id: str) -> list[Artifact]:
        if not self.get_task(run_id, owner_id):
            return []
        path = self._artifacts_path(run_id)
        if not path.exists():
            return []
        return [Artifact.model_validate(item) for item in _read_json_list(path)]

    def list_runs(self, owner_id: str, limit: int = 20, offset: int = 0) -> list[RunTask]:
        tasks = []
        for run_dir in self.root.iterdir():
            if not run_dir.is_dir():
                continue
            task_path = run_dir / "run.json"
            if not task_path.exists():
                continue
            try:
                data = _read_json_text(task_path)
                if data is None:
                    continue
                task = RunTask.model_validate_json(data)
                if task.owner_id == owner_id:
                    tasks.append(task)
            except Exception:
                continue
        tasks.sort(key=lambda t: t.created_at or datetime.min, reverse=True)
        return tasks[offset : offset + limit]

    def list_runs_for_cleanup(self, cutoff: datetime, limit: int = 100000) -> list[RunTask]:
        tasks = []
        for run_dir in self.root.iterdir():
            if not run_dir.is_dir():
                continue
            task_path = run_dir / "run.json"
            if not task_path.exists():
                continue
            try:
                data = _read_json_text(task_path)
                if data is None:
                    continue
                task = RunTask.model_validate_json(data)
                if task.created_at and task.created_at < cutoff:
                    tasks.append(task)
            except Exception:
                continue
        tasks.sort(key=lambda t: t.created_at or datetime.min)
        return tasks[:limit]

    def delete_run(self, run_id: str, owner_id: str) -> bool:
        task = self.get_task(run_id, owner_id)
        if not task:
            return False
        self._remove_run_files(run_id)
        shutil.rmtree(self._run_dir(run_id), ignore_errors=True)
        return True

    def delete_runs(self, owner_id: str) -> int:
        run_ids = [task.run_id for task in self.list_runs(owner_id, limit=100000, offset=0)]
        deleted = 0
        for run_id in run_ids:
            if self.delete_run(run_id, owner_id):
                deleted += 1
        return deleted

    def _run_dir(self, run_id: str) -> Path:
        return self.root / run_id

    def _task_path(self, run_id: str) -> Path:
        return self._run_dir(run_id) / "run.json"

    def _events_path(self, run_id: str) -> Path:
        return self._run_dir(run_id) / "events.json"

    def _artifacts_path(self, run_id: str) -> Path:
        return self._run_dir(run_id) / "artifacts.json"

    def _remove_run_files(self, run_id: str) -> None:
        uploads_dir = _abs_path(self.settings.storage.uploads_dir) / run_id
        reports_dir = _abs_path(self.settings.storage.reports_dir) / run_id
        shutil.rmtree(uploads_dir, ignore_errors=True)
        shutil.rmtree(reports_dir, ignore_errors=True)


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{uuid4().hex}.tmp")
    tmp_path.write_text(text, encoding="utf-8")
    tmp_path.replace(path)


def _read_json_text(path: Path, attempts: int = 3) -> str | None:
    for attempt in range(attempts):
        try:
            text = path.read_text(encoding="utf-8")
        except OSError:
            return None
        if text.strip():
            return text
        if attempt < attempts - 1:
            time.sleep(0.02)
    return None


def _read_json_list(path: Path) -> list[dict]:
    text = _read_json_text(path)
    if text is None:
        return []
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return []
    return data if isinstance(data, list) else []


def create_text_input(run_id: str, text: str, settings: Settings) -> Path:
    upload_dir = _abs_path(settings.storage.uploads_dir) / run_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    path = upload_dir / "input.md"
    path.write_text(text, encoding="utf-8")
    return path


def create_file_input(run_id: str, filename: str, content: bytes, settings: Settings) -> Path:
    upload_dir = _abs_path(settings.storage.uploads_dir) / run_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(filename).name or "input"
    path = upload_dir / safe_name
    path.write_bytes(content)
    return path


def execute_run(run_id: str, owner_id: str, store: TaskStore | None = None) -> RunTask:
    store = store or LocalTaskStore()
    task = store.get_task(run_id, owner_id)
    if task is None:
        raise KeyError(f"run not found: {run_id}")

    logger.info(f"execute_run: starting task {run_id}")
    usage_accumulator = UsageAccumulator()
    usage_lock = threading.Lock()

    def persist_usage(entry: dict) -> None:
        usage_summary = compact_usage_summary(usage_accumulator.add(entry))
        global_usage = add_persistent_usage(store.settings, entry)
        with usage_lock:
            latest = store.get_task(run_id, owner_id) or task
            latest.summary = {
                **(latest.summary or {}),
                "usage": usage_summary,
                "global_usage": global_usage,
                "cost_usd": usage_summary["estimated_cost_usd"],
            }
            store.save_task(latest)

    original_usage_callback = store.settings.usage_callback
    store.settings.usage_callback = persist_usage

    try:
        logger.info(f"execute_run: updating status to running for {run_id}")
        task.status = "running"
        task.current_stage = "ingest"
        task.started_at = datetime.now()
        task.summary = {
            **(task.summary or {}),
            "usage": compact_usage_summary(usage_accumulator.snapshot()),
            "global_usage": read_persistent_usage(store.settings),
            "cost_usd": 0.0,
        }
        store.save_task(task)
        store.add_event(run_id, "running", "info", "Run started")
        logger.info(f"execute_run: status saved as running for {run_id}")

        output_dir = _abs_path(store.settings.storage.reports_dir) / run_id
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"result.{task.requested_format}"
        logger.info(f"execute_run: starting orchestrator for {run_id}")
        report = run(
            task.input_path or "",
            out_path=output_path,
            fmt=task.requested_format,
            config=store.settings,
            detailed=task.detailed,
        )
        logger.info(f"execute_run: orchestrator completed for {run_id}")

        artifact = artifact_from_file(output_path, task.run_id, task.owner_id, task.requested_format)
        store.save_artifact(artifact)
        artifacts = [artifact]

        for extra_fmt in _extra_output_formats(task.requested_format):
            extra_path = output_dir / f"result.{extra_fmt}"
            if _write_extra_artifact(output_path, extra_path, task.requested_format, extra_fmt):
                extra_artifact = artifact_from_file(extra_path, task.run_id, task.owner_id, extra_fmt)
                store.save_artifact(extra_artifact)
                artifacts.append(extra_artifact)

        logger.info(f"execute_run: marking as completed for {run_id}")
        task.status = "completed"
        task.current_stage = "completed"
        task.finished_at = datetime.now()
        usage_summary = compact_usage_summary(usage_accumulator.snapshot())
        task.summary = {
            **report.summary,
            "usage": usage_summary,
            "global_usage": read_persistent_usage(store.settings),
            "cost_usd": usage_summary["estimated_cost_usd"],
        }
        task.total_claims = int(report.summary.get("total", 0) or 0)
        task.completed_claims = task.total_claims
        task.artifact_ids = [item.artifact_id for item in artifacts]
        store.save_task(task)
        store.add_event(run_id, "completed", "info", "Run completed", task.completed_claims, task.total_claims)
        logger.info(f"execute_run: completed successfully for {run_id}")
        return task
    except Exception as exc:
        logger.error(f"execute_run: error for {run_id}: {exc}", exc_info=True)
        task.status = "failed"
        task.current_stage = "failed"
        task.error = str(exc)
        task.finished_at = datetime.now()
        store.save_task(task)
        store.add_event(run_id, "failed", "error", str(exc))
        logger.info(f"execute_run: marked as failed for {run_id}")
        return task
    finally:
        store.settings.usage_callback = original_usage_callback


def artifact_from_file(path: Path, run_id: str, owner_id: str, fmt: str) -> Artifact:
    body = path.read_bytes()
    content_type = CONTENT_TYPES.get(fmt) or mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return Artifact(
        artifact_id=str(uuid4()),
        run_id=run_id,
        owner_id=owner_id,
        fmt=fmt,  # type: ignore[arg-type]
        storage_uri=path.resolve().as_uri(),
        filename=path.name,
        content_type=content_type,
        size_bytes=len(body),
        sha256=hashlib.sha256(body).hexdigest(),
        created_at=datetime.now(),
    )


def _extra_output_formats(requested_format: str) -> list[str]:
    if requested_format == "json":
        return ["xlsx"]
    return []


def _write_extra_artifact(source_path: Path, target_path: Path, source_fmt: str, target_fmt: str) -> bool:
    if source_fmt == "json" and target_fmt == "xlsx":
        _write_xlsx_from_json_report(source_path, target_path)
        return True
    return False


def _write_xlsx_from_json_report(source_path: Path, target_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("xlsx output requires openpyxl") from exc

    data = json.loads(source_path.read_text(encoding="utf-8"))
    sections = data if isinstance(data, list) else []
    wb = Workbook()
    ws = wb.active
    ws.title = "source_verification"

    if sections and all(isinstance(item, dict) and "rows" in item for item in sections):
        for section in sections:
            title = str(section.get("title") or "sub-table")
            headers = list(section.get("headers") or [])
            rows = list(section.get("rows") or [])
            ws.append([title])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for row in rows:
                ws.append([row.get(header, "") if isinstance(row, dict) else "" for header in headers])
            ws.append([])
    else:
        rows = sections
        headers = list(rows[0].keys()) if rows and isinstance(rows[0], dict) else []
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(header, "") if isinstance(row, dict) else "" for header in headers])

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 24

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)


def path_from_storage_uri(storage_uri: str) -> Path:
    if storage_uri.startswith("file:///"):
        from urllib.parse import unquote, urlparse

        parsed = urlparse(storage_uri)
        path = unquote(parsed.path)
        if parsed.netloc:
            return Path(f"//{parsed.netloc}{path}")
        if len(path) >= 4 and path[0] == "/" and path[2] == ":":
            path = path[1:]
        return Path(path)
    return Path(storage_uri)


def owner_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _abs_path(path: str) -> Path:
    candidate = Path(path)
    return candidate if candidate.is_absolute() else ROOT / candidate
